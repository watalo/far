#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@NAME		:dataio.py
@TIME		:2022/06/29 14:26:01
@AUTHOR     :watalo
@VERSION	:0.0.x

模块内容：
对原始财务报表进行读取、格式化
    - 读取: pd.read_excel --> df
    - 格式化:
        - 列: 
            - 科目：增加各种财务指标，需要进行财务科目进行计算
            - 汇总：增加仅需要汇总相加的指标，如刚性兑付、营运资金量等指标
        - 行: 增加每年比上年的增量、增幅、在总资产、总负债、总权益、总收入占比
'''

from lib2to3.pgen2.pgen import DFAState
import os
import pandas as pd
from openpyxl import load_workbook
from resourcepath import *


def read_fs_data(input_path):
    '''
    读取xlsx文件数据，为避免xlsx文件中包含公式，需要进行处理。
    @input_path: 原始报表数据的路径，相对/绝对都可
    '''
    # 使用openpyxl.load_workbook的data_only参数进行强制转换
    ws = load_workbook(input_path,
                       data_only=True)
    ws.save(os.path.abspath(input_path))
    fs_df = pd.read_excel(io=input_path,
                          sheet_name='Sheet1',
                          index_col=0)  # 强制将科目名称作为行标签
    # 报表通常格式-->行：科目名称，列：日期
    return fs_df  # 不知道需不需要转置，先这么滴


def format_df(fs_df):
    '''
    1.设置科目属性：
    新增列: name="科目分类"
    新增列: name="流动性", [False, True, NaN]
    新增列: name= 前3年d、前2年d、前1年d、当期d
        从而实现
        - 资产
            - 流动资产
            - 非流动资产
        - 负债
            - 流动负债
            - 非流动负债
        - 权益
        - 损益
        - 现金流
    2.设置财务指标和预先计算数据
    新增行：    
        - 财务指标
            - 营运指标
            - 流动性分析
            - 偿债能力 
        - 统计统计
            - sum
            - mean
    '''
    
    columns = fs_df.columns.tolist()

    # 科目明细列表
    rows = fs_df.index.tolist()
    asset_list  = rows[ :34]
    debt_list   = rows[34:58]
    equity_list = rows[58:66]
    gain_list   = rows[66:86]
    cashflow_list = rows[86:]

    add_columns = ['前2年d', '前1年d', '当期d',
                   '前2年r', '前1年r', '当期r',
                   '前3年p','前2年p', '前1年p', '当期p',
                   '科目分类', '流动性']


    add_rows = ['资产负债率', '流动比率', '速动比率']
    fs_df = fs_df.reindex(index=fs_df.index.tolist() + add_rows,
                          columns=fs_df.columns.tolist() + add_columns
                          )
    # 计算增量
    for index, i in enumerate(['前2年d', '前1年d', '当期d']):
        fs_df[i] = fs_df.iloc[:, index+1] - fs_df.iloc[:, index]
    # 计算增幅
    for index, i in enumerate(['前2年r', '前1年r', '当期r']):
        fs_df[i] = fs_df.iloc[:, index+1]/fs_df.iloc[:, index] - 1
    # 计算占比
    for index, i in enumerate(['前3年p','前2年p', '前1年p', '当期p']):
        fs_df[i] = fs_df.iloc[:, index]/fs_df.iloc[rows.index('资产总计'), index]
    # 科目类别赋值
    for i in rows:
        if i in asset_list:
            fs_df.loc[i,'科目分类'] = '资产'
        elif i in debt_list:
            fs_df.loc[i,'科目分类'] = '负债'
        elif i in equity_list:
            fs_df.loc[i,'科目分类'] = '权益'
        elif i in gain_list:
            fs_df.loc[i,'科目分类'] = '损益'
        elif i in cashflow_list:
            fs_df.loc[i,'科目分类'] = '现金流'


    print(rows)
    print(asset_list)
    print(debt_list)
    print(equity_list)
    print(gain_list)
    print(cashflow_list)
    return fs_df


if __name__ == '__main__':

    input_path = './far/far/test_financial_statement.xlsx'

    fs_data = read_fs_data(input_path)
    # add_row = ['前2年d', '前1年d', '当期d', '科目分类', '流动性']
    # add_row = ['资产负债率', '流动比率', '速动比率']
    # fs_data = fs_data.reindex(index=fs_data.index.tolist() + add_row,
    #                           columns=fs_data.columns.tolist() + add_colmuns
    #                           )
    new = format_df(fs_df=fs_data)
    print(new.iloc[:, 4:])
